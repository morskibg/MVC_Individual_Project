from math import log, ceil, frexp
import pandas as pd
from decimal import Decimal
import os, sys

import barcode
from barcode.writer import ImageWriter

from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY, TA_RIGHT
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Frame, KeepInFrame
from reportlab.platypus import Flowable, Indenter, Table, TableStyle
from reportlab.lib.units import inch, mm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.fonts import addMapping
from reportlab.graphics import shapes, renderPM
from reportlab.graphics.shapes import Drawing
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode import eanbc
from reportlab.lib import colors
import textwrap

import xlsxwriter
from openpyxl.styles import Alignment
from openpyxl import load_workbook

from app import app
from app.models import *
from app.helpers.helper_functions import convert_date_to_utc, update_or_insert



class StaticData:

    def __init__(self):

        self.bank_data = \
                "БАНКА : <b>ДСК БАНК ЕАД</b><br />\
                СМЕТКА : <b>BGL-BG41STSA93000021750244 / STSABGSF</b> <br />\
                БАНКА : <b>ТБ ИНВЕСТБАНК ЕАД</b> <br />\
                СМЕТКА : <b>BGL- BG71IORT81111003823100 / IORTBGSF</b> <br />\
                БАНКА : <b>ТБ ЮРОБАНК ЕАД</b> <br />\
                СМЕТКА : <b>BGL-BG77BPBI79401087902201 / BPBIBGSF </b><br />\
                БАНКА : <b>УНИКРЕДИТ БУЛБАНК АД </b><br />\
                СМЕТКА : <b>BGL-BG91UNCR70001521763920 / UNCRBGSF</b> <br />"

        
            

    def get_bank_data(self):
        return self.bank_data



class ParagraphStyler(ParagraphStyle):

    def __init__(self):       
        self.__dict__.update(self.defaults)
        self.alignment  = TA_LEFT
        self.leading = 15
        self.fontName = 'DejaVuSerif'
        self.fontSize = 12
        self.leftIndent = 0
        self.rightIndent = 0
        
    def medium_text(self):
        self.alignment  = TA_LEFT
        self.leading = 15
        self.fontName = 'DejaVuSerif'
        self.fontSize = 10

    def reset(self):
        self.__init__()


class SumTextWriter:

    def __init__(self):
        self.num = None

    def __under_20__(self, is_stotinki = False):
        
        num_dict =  {0:'нула',1:'един', 2:'два',3:'три',4:'четири',5:'пет',6:'шест',7:'седем',8:'осем',9:'девет',10:'десет',11:'единайсет', \
                    12:'дванайсет', 13:'тринайсет',14:'четиринайсет',15:'петнайсет',16:'шестнайсет',17:'седемнайсет',18:'осемнайсет',19:'деветнайсет'}
        
        if self.num in range(0,20):
            res = num_dict[self.num]
            if is_stotinki:
                if self.num == 1:
                    res = 'една'
                elif self.num == 2:
                    res = 'две'
                            
            return res
        else:                      
            return None

    def __between_20_100__(self, is_stotinki = False):
        
        if self.num < 20:
            return self.__under_20__(is_stotinki)
        else:
            num_dict = {2:'двайсет', 3:'трийсет',4:'четирийсет',5:'петдест',6:'шестдест',7:'седемдесет',8:'осемдесет',9:'деветдесет'}
            tenths = self.num // 10
            ones = self.num - tenths * 10
            word = str(num_dict[tenths])
            if ones != 0:
                self.num = ones                
                word += f' и {self.__under_20__(is_stotinki)}'
            return word
        
    def __between_100_1000__(self):
        
        if self.num < 100:
            return self.__between_20_100__()
        else:
            
            num_dict = {1:'сто',2:'двеста', 3:'триста',4:'четиристотин',5:'петстотин',6:'шестотин',7:'седемдестотин',8:'осемстотин',9:'деветстотин'}
            hundreds = self.num // 100
            tenths = self.num - hundreds * 100
            word = f'{num_dict[hundreds]}'
            self.num = tenths            
            addition_words = self.__between_20_100__()
                       
            if tenths != 0:
                word += f' и {addition_words}' if len(addition_words) < 13 and len(addition_words.split(' ')) <= 1 else f' {addition_words}'
           
            return word   
        
    def __between_1000_1000000__(self):
        
        initial_num = self.num
        if self.num < 1000:
            return self.__between_100_1000__()

        elif self.num < 2000:
            if self.num == 1000:
                return 'хиляда'
            else:
                self.num -= 1000
                       
                words = self.__between_100_1000__()                
                return f'хиляда и {words}' if len(words) < 13 and len(words.split(' ')) <= 1 else f'хиляда {words}'
        else:
            
            num_dict =  { 2:'двe',3:'три',4:'четири',5:'пет',6:'шест',7:'седем',8:'осем',9:'девет',10:'десет',11:'единайсет', 12:'дванайсет', \
                        13:'тринайсет',14:'четиринайсет',15:'петнайсет',16:'шестнайсет',17:'седемнайсет',18:'осемнайсет',19:'деветнайсет'}
            thousands = self.num // 1000
            
            self.num = thousands
            
            if thousands < 20:
                first_word = num_dict[thousands]
            elif thousands < 100:
                first_word = self.__between_20_100__()
            elif thousands < 1000:
                first_word = self.__between_100_1000__() 
            
            hundreds = initial_num - thousands * 1000
            self.num = hundreds
            if hundreds == 0:
                return  f'{first_word} хиляди'
            
            addition_words = self.__between_100_1000__()
            words = f'{first_word} хиляди и {addition_words}' if len(addition_words) < 13 and len(addition_words.split(' ')) <= 1 else f'{first_word} хиляди {addition_words}'
            return words


    def __over_milion__(self):
        if self.num < 1000000:
            return self.__between_1000_1000000__()
        else:
            milions = self.num // 1000000
            thousands = self.num - milions * 1000000

            self.num = milions
            if milions < 20:                
                first_word = self.__under_20__()
            else:
                first_word = self.__between_20_100__(milions)

            self.num = thousands    
            if thousands == 0:
                if milions == 1:
                    words = f'{first_word} милион'
                else:    
                    words = f'{first_word} милиона '
            else:
                if milions == 1:
                     words = f'{first_word} милион {self.__between_1000_1000000__()}'                
                else:
                    words = f'{first_word} милиона {self.__between_1000_1000000__()}'

            return(words)

    def number_to_words(self, ext_num):
        # print(f'from word creation EXT_NUM --------->{ext_num}')
        self.num = int(ext_num)
        integer_str = self.__over_milion__().capitalize()
        self.num = int(round((ext_num - int(ext_num))*100))
        if int(round((ext_num - int(ext_num))*100)) == 100:
            self.num = 0
            decimal_str = '0'
            self.num = int(ext_num) + 1
            integer_str = self.__over_milion__().capitalize()
        else:
            decimal_str = self.__between_20_100__(is_stotinki = True)


        # print(f'from word creation --------->{int(ext_num)} ______ {self.num}')       
        
        
        return f'{integer_str} лева и {decimal_str} стотинки.'



class InvoiceCreator:

    def __init__(self, pdf_file_name, df, p_style, sum_writer):
        
        self.raw_df = df
        self.p_style = p_style
        self.sum_writer = sum_writer
        self.canvas = canvas.Canvas(os.path.join(os.path.join(app.root_path, app.config['PDF_INVOICES_PATH']), pdf_file_name), pagesize=A4)
        self.styles = getSampleStyleSheet()
        self.width, self.height = letter
        self.grid = df[df['StockName'] == 'ПРЕНОС И ДОСТЪП ДО ЕЛ.МРЕЖАТА']
        self.grid_tuple = self.__create_table_tuple__(self.grid)
        self.zko = df[df['StockName'] == 'ЦЕНА ЗАДЪЛЖЕНИЕ КЪМ ОБЩЕСТВОТО СЪГЛАСНО ЧЛ.100 АЛ.4 ОТ ЗЕ И ЧЛ.31 ОТ']
        self.zko_tuple = self.__create_table_tuple__(self.zko)
        self.akciz = df[df['StockName'] == 'НАЧИСЛЕН АКЦИЗ']
        self.akciz_tuple = self.__create_table_tuple__(self.akciz)
        self.power = df[df['StockName'] == 'ELECTRICITY - (ЕЛЕКТРИЧЕСКА ЕНЕРГИЯ)']
        self.power_tuple = self.__create_table_tuple__(self.power)   
        self.lead_data = None
        self.__specify_lead_data__()
        self.y_offset = 10
        self.total_sum = 0
        self.vat = 0
        self.net_sum = 0

    def __specify_lead_data__(self):

        if self.akciz is not None:
            self.lead_data = self.akciz
        elif self.zko is not None:
            self.lead_data = self.zko
        elif self.grid is not None:
            self.lead_data = self.grid
        else:
            self.lead_data = self.power  



    def coord(self, x, y, unit=1):
        x, y = x * unit, self.height - y * unit
        return x, y
    
    def __space_gen__(self,count):
        return '&nbsp;' * count
    
    def __paragraph_writer__(self, p_text, ext_x, ext_y, style = None):

        style = self.styles["Normal"] if style is not None else self.p_style

        p = Paragraph(p_text, style)
        p.wrapOn(self.canvas, self.width, self.height)
        p.drawOn(self.canvas, *self.coord(ext_x, ext_y, mm))

    def __generate_full_invoice_number__(self):

        SIZE = 10
        part_num = int(self.lead_data['DocNumber'].values[0])        
        count = SIZE - int(ceil(log(part_num + 1, 10)))
        full_num = '0' * count + str(part_num)
        return full_num

    def __create_table_tuple__(self, df):
        if df.empty:
            qty = 0
            price = amount = amount_calc = '0.00'
        else:
            qty = f'{df.Quantity.values[0]:,.3f}'.replace(',', u'\u2009') 
            price = f'{df.PriceLev.values[0]:,.2f}'.replace(',', u'\u2009')
            amount = f'{df.ItemSuma.values[0]:,.2f}'.replace(',', u'\u2009')            
            amount_calc = f'{df.ItemSuma.values[0]:.2f}'
            
        return(qty, price, amount, amount_calc)        

    def create_barcode(self, ext_x = 8 , ext_y = 20):
    
        ext_y -=  self.y_offset
        data = '28' +  self.__generate_full_invoice_number__()
        barcode = eanbc.Ean13BarcodeWidget(data)
        bounds = barcode.getBounds()
        w = float(130)
        h = float(50)
        bar_width = bounds[2] - bounds[0]
        bar_height = bounds[3] - bounds[1]
        d = Drawing(w, h, transform=[w / bar_width, 0, 0, h / bar_height, 0, 0])
        d.add(barcode)
        renderPDF.draw(d, self.canvas, *self.coord(ext_x, ext_y, mm))

    def create_invoice_number(self, ext_x = 70 , ext_y = 13):

        ext_y -=  self.y_offset

        curr_num = self.__generate_full_invoice_number__()
        curr_date = pd.to_datetime(self.lead_data['DocDate'].values[0], format='%d.%m.%Y') 
        p_text = f'<b>ФАКТУРА ОРИГИНАЛ</b><br />'
        p_text += f'<b>{curr_num} / {curr_date.date()}</b>'
        
        self.__paragraph_writer__(p_text, ext_x, ext_y)
           
           
    def create_logo(self, ext_x = 140 , ext_y = 25):

        ext_y -=  self.y_offset
        logo = 'app/static/img/Grand_Energy_logo_.png'
        self.canvas.drawInlineImage(logo,*self.coord(ext_x, ext_y, mm),180, 90)

    def create_seller_data(self, ext_x = 107 , ext_y = 50):

        ext_y -=  self.y_offset
        name = self.lead_data['SNAME'].values[0]
        ein = self.lead_data['STAXNUM'].values[0]
        vat = self.lead_data['SBULSTAT'].values[0]
        address = self.lead_data['SAddress'].values[0]
        store_name = self.lead_data['StoreName'].values[0]  

        p_text =  f'ДОСТАВЧИК : <b>{name}</b><br />' 
        p_text += f'ЕИН : <b>{ein}</b><br />'        
        p_text += f'ДДС No    : <b>{vat}</b><br />'
        p_text += f'АДРЕС     : <b>{address}</b><br />'
        # p_text += f'ОБЕКТ     : <b>{store_name}</b><br />'

        self.p_style.leading = 12
        self.p_style.fontSize = 8
        self.__paragraph_writer__(p_text, ext_x, ext_y)

    def create_buyer_data(self, ext_x = 10 , ext_y = 50):
        
        name = self.lead_data['FirmName'].values[0]
        ext_y -=  self.y_offset
        p_text =  'ПОЛУЧАТЕЛ : <b>'
        wrap_text = textwrap.wrap(name, width=30)
        for idx,p in enumerate(wrap_text, start=1):
            
            if idx == 1:
                p_text += f'{p}</b><br />'  
                           
            elif idx == 2: 
                ext_y += 4
                p_text += f'{self.__space_gen__(5)}<b>{p}</b><br />'  
            else:
                break

        ein = self.lead_data['TaxNum'].values[0]
        vat = self.lead_data['BULSTAT'].values[0] if not pd.isnull(self.lead_data['BULSTAT'].values[0]) else ''
        store_name = ''      
        city_ = self.lead_data['CityName'].values[0] if isinstance(self.lead_data['CityName'].values[0], str) else ''   
        address_ = self.lead_data['Address'].values[0] if isinstance(self.lead_data['Address'].values[0], str) else ''

        address = city_ + ', ' + address_

        p_text += f'ЕИН : <b>{ein}</b><br />'        
        p_text += f'ДДС No    : <b>{vat}</b><br />'
        p_text += 'АДРЕС     : <b>'
        wrap_text = textwrap.wrap(address, width=30)
        for idx,p in enumerate(wrap_text, start=1):
            
            if idx == 1:
                p_text += f'{p}</b><br />'  
                           
            elif idx == 2:
                ext_y += 4
                p_text += f'{self.__space_gen__(5)}<b>{p}</b><br />'  

            else:
                break       
        
        # p_text += f'ОБЕКТ     : <b>{store_name}</b><br />'

        self.p_style.leading = 12
        self.p_style.fontSize = 8
        self.__paragraph_writer__(p_text, ext_x, ext_y )

    def create_table(self,ext_x = 10, ext_y = 103):

        font_size = 8

        data1 = [['№\n ', 'Наименование на стоките и услугите\n ', 'Количество\nМВтч', 'Ед.Цена\nлв/МВтч', 'Стойност\nлева'],
                ['1',  'Електрическа енергия\n(ELECTRICITY)',self.power_tuple[0], self.power_tuple[1], self.power_tuple[2]],
                ['2',  'Задължение към обществото \nсъгласно ЧЛ.100, АЛ.4 и ЧЛ.31 от ЗЕ', self.zko_tuple[0], self.zko_tuple[1], self.zko_tuple[2]],
                ['3',   'Начислен акциз\nсъгласно ЧЛ.20, АЛ.2,Т.17 от ЗАДС', self.akciz_tuple[0], self.akciz_tuple[1], self.akciz_tuple[2]],
                ['4', 'Пренос и достъп \nдо електрическата мрежа.', self.grid_tuple[0], self.grid_tuple[1], self.grid_tuple[2]]]

        t1 = Table(data1)
        
        t1.setStyle(TableStyle([
            ('FONT', (0, 0), (4, 0),'DejaVuSerifBold'),
            ('FONT', (0, 1), (4, 1),'DejaVuSerif'),
            ('FONT', (0, 2), (4, 2),'DejaVuSerif'),
            ('FONT', (0, 3), (4, 3),'DejaVuSerif'),
            ('FONT', (0, 4), (4, 4),'DejaVuSerif'),

            ('FONTSIZE', (0, 1), (4, 1),font_size),
            ('FONTSIZE', (0, 2), (4, 2),font_size),
            ('FONTSIZE', (0, 3), (4, 3),font_size),
            ('FONTSIZE', (0, 4), (4, 4),font_size),
            
            ('ALIGN', (0, 0), (4, 0), "CENTER"),
            ('ALIGN', (0, 1), (4, 1), "CENTER"),
            ('ALIGN', (0, 2), (4, 2), "CENTER"),
            ('ALIGN', (0, 3), (4, 3), "CENTER"),
            ('ALIGN', (0, 4), (4, 4), "CENTER"),
            # ('BACKGROUND', (0, 0), (4, 0), colors.HexColor("#F48120")),
            ('BACKGROUND', (0, 0), (4, 0), colors.HexColor("#dfdfdf")),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),

        ]))
        t1.wrapOn(self.canvas, 100, 100)
        t1.drawOn(self.canvas,*self.coord(ext_x, ext_y, mm))


    def create_bank_data(self, ext_x = 10, ext_y = 205):
        # ext_y += self.y_offset

        font_size = 7
        font_size_header = 8
        self.p_style.alignment= TA_LEFT
        bank_table_data=[['№','Банка','Сметка (IBAN)','BIC Code', 'Валута'],
                    ['1','ТБ ИНВЕСТБАНК ЕАД','BG71IORT81111003823100                ','IORTBGSF','лева'],
                    ['2','УНИКРЕДИТ БУЛБАНК АД','BG91UNCR70001521763920','UNCRBGSF','лева'],
                    ['3','ТБ ЮРОБАНК ЕАД','BG77BPBI79401087902201','BPBIBGSF','лева'],
                    ['4','ДСК БАНК ЕАД','BG41STSA93000021750244','STSABGSF','лева']]

        bank_table = Table(bank_table_data)
        bank_table.setStyle(TableStyle([
            ('FONT', (0, 0), (4, 0),'DejaVuSerif'),
            ('FONT', (0, 1), (4, 1),'DejaVuSerif'),
            ('FONT', (0, 2), (4, 2),'DejaVuSerif'),
            ('FONT', (0, 3), (4, 3),'DejaVuSerif'),
            ('FONT', (0, 4), (4, 4),'DejaVuSerif'),
            ('FONTSIZE', (0, 0), (4, 0),font_size_header),
            ('FONTSIZE', (0, 1), (4, 1),font_size),
            ('FONTSIZE', (0, 2), (4, 2),font_size),
            ('FONTSIZE', (0, 3), (4, 3),font_size),
            ('FONTSIZE', (0, 4), (4, 4),font_size),
            # ('BACKGROUND', (0, 0), (4, 0), colors.orange),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.black),
        ]))

        bank_table.wrapOn(self.canvas, 100, 100)
        bank_table.drawOn(self.canvas,*self.coord(ext_x, ext_y, mm))

        if self.raw_df.iloc[0].EasyPay == 1: #!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
            self.p_style.fontSize = 8
            self.p_style.leading = 11
            easy_pay_custom_num = int(self.raw_df.iloc[0].easy_pay_num)
            easy_pay_name = self.raw_df.iloc[0].easy_pay_name
            p_text = f'<b>За плащане през EasyPay / ePay, Вашите \'Име\' - \'Клиентски номер\' са:</b><br /> <b>{easy_pay_name} - {easy_pay_custom_num}</b>' 
            self.__paragraph_writer__(p_text, ext_x, ext_y + 9)
        
    
    def create_payment_summary(self, ext_x = 10 , ext_y = 115):

        # ext_y += self.y_offset

        sum_neto = Decimal(str(self.grid_tuple[3])) + Decimal(str(self.zko_tuple[3])) + \
                     Decimal(str(self.akciz_tuple[3])) + Decimal(str(self.power_tuple[3]))

        self.net_sum = sum_neto

        sum_vat = Decimal(str(self.raw_df.iloc[0].VATValue)) * Decimal('0.01') * Decimal(str(sum_neto))
        self.vat = sum_vat

        total_sum = Decimal(str(sum_neto)) + Decimal(str(sum_vat))
        self.total_sum = total_sum

        no_vat_reason = self.lead_data['VATDescrBG'].values[0]

        shft_y = 8
        self.p_style.fontSize = 8
        p_text = '<b>Общо сума</b>'
        self.__paragraph_writer__(p_text, ext_x, ext_y - 2)
        self.canvas.line(*self.coord(ext_x, ext_y , mm),*self.coord(ext_x + 182, ext_y , mm) )   

        p_text = 'Данъчна основа<br/>'
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)
        self.canvas.setLineWidth(0.25)
        self.canvas.line(*self.coord(ext_x, ext_y + 1 + shft_y, mm),*self.coord(ext_x + 182, ext_y + 1 + shft_y, mm) )

        p_text = f'ДДС на данъчна основа: (Ставка на ДДС: {self.raw_df.iloc[0].VATValue}%)<br/>'
        shft_y += 8
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)        
        self.canvas.line(*self.coord(ext_x, ext_y + 1 + shft_y, mm),*self.coord(ext_x + 182, ext_y + 1 + shft_y, mm) )

        p_text = f'Основание за нулева ставка / неначисляване на ДДС: {no_vat_reason}<br/>'
        shft_y += 8
        self.canvas.setLineWidth(1.75)
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)        
        self.canvas.line(*self.coord(ext_x, ext_y + 1 + shft_y, mm),*self.coord(ext_x + 182, ext_y + 1 + shft_y, mm) )

        p_text = f'<b>Сума за плащане</b><br/>'
        # self.canvas.setLineWidth(1.75)
        self.p_style.fontSize = 10
        shft_y += 8
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)        
        # self.canvas.line(*self.coord(ext_x, ext_y + 1 + shft_y, mm),*self.coord(ext_x + 182, ext_y + 1 + shft_y, mm) )
        

        self.p_style.fontSize = 7
        shft_y += 7
        p_text = f'Словом сума за плащане: {self.sum_writer.number_to_words(total_sum)}'
        wrap_text = textwrap.wrap(p_text, width=120)
        for idx,p in enumerate(wrap_text, start=1):
            if idx != 1:
                self.p_style.leftIndent = 100
                p = f'{p}<br/>'
            else:                
                p += '<br/>'                
            
            self.__paragraph_writer__(p, ext_x, ext_y + shft_y)
            shft_y += 3
        self.p_style.leftIndent = 0
        # shft_y += 2
        p_text = 'Основание &nbsp;за &nbsp;начисляване &nbsp;на &nbsp;акциз &nbsp;чл.20,ал.2,т.17  \
            &nbsp;от &nbsp;ЗАДС;ИН:BG005800S0071;&nbsp;Код &nbsp;по &nbsp;АП:E012;&nbsp;Код &nbsp;по &nbsp;КН:271600000<br/>'
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)

        
        self.p_style.medium_text()
        # self.p_style.fontSize = 10
        shft_y += 8
        
        curr_date = pd.to_datetime(self.lead_data['ReceiveDate'].values[0], format='%d.%m.%Y')
        p_text = 'Начин на плащане: по банков път<br/>'
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)
        shft_y += 5
        self.p_style.fontSize = 7
        p_text = f'Дата на данъчно събитие: {curr_date.date()}'
        self.__paragraph_writer__(p_text, ext_x, ext_y + shft_y)


        self.p_style.alignment = TA_RIGHT
        shft_y = 0
        ext_x -= 5
        p_text = f'<b>{sum_neto:,.2f}'.replace(',', u'\u2009') +' лв.</b><br/>'
        self.__paragraph_writer__(p_text, ext_x - 30, ext_y + shft_y)

        shft_y += 8
        p_text = f'{sum_neto:,.2f}'.replace(',', u'\u2009') +' лв.<br/>'
        self.__paragraph_writer__(p_text, ext_x - 30, ext_y + shft_y)

        shft_y += 8
        p_text = f'{sum_vat:,.2f}'.replace(',', u'\u2009') +' лв.<br/>'
        self.__paragraph_writer__(p_text, ext_x - 30, ext_y + shft_y)

        shft_y += 8
        p_text = f'{no_vat_reason}<br />'
        self.__paragraph_writer__(p_text, ext_x - 30, ext_y + shft_y)

        shft_y += 8
        self.p_style.fontSize = 10
        ext_x += 1
        p_text = f'<b>{total_sum:,.2f}'.replace(',', u'\u2009') +' лв.</b><br/>'
        self.__paragraph_writer__(p_text, ext_x - 30, ext_y + shft_y)

        self.p_style.reset()

    def create_signature_fields(self, ext_x = 10 , ext_y = 225 ):

        # ext_y += self.y_offset

        self.p_style.medium_text()
        pay_date = pd.to_datetime(self.lead_data['PayDate'].values[0], format='%d.%m.%Y')
        # self.p_style.fontSize = 10
        p_text = f'Дата на падеж: {pay_date.date()}<br />'
        p_text += 'Приел :'
        # print(f'{p_text}')
        self.__paragraph_writer__(p_text, ext_x, ext_y)
        
        self.p_style.alignment = TA_RIGHT
        ext_x = -25
        p_text = f'Съставил: ID  {self.lead_data.CreatorID.values[0]} {self.lead_data.PersonName.values[0]}<br/>'
        self.__paragraph_writer__(p_text, ext_x, ext_y)
        self.p_style.medium_text()
    
    def create_notes(self, ext_x = 10 , ext_y = 235):

        ext_y -= self.y_offset/3
        y_shift = 3
        self.p_style.fontSize = 6
        self.canvas.setLineWidth(0.5)
        # self.leading = 3
        p_text = 'Забележки :<br/>'
        self.canvas.line(*self.coord(ext_x, ext_y -2 , mm),*self.coord(ext_x + 182, ext_y - 2, mm) )
        self.__paragraph_writer__(p_text, ext_x, ext_y)
        # ext_y += self.y_offset/2
        self.p_style.leading = 7
        p_text = 'Настоящата фактура се издава без подпис и печат. Тази промяна е съгласно чл.8 от ЗС и чл. 78, ал. 8 от ППЗДДС.<br/>'
        p_text += 'Документът е подписан с Универсален електронен подпис.<br/>'        
        self.__paragraph_writer__(p_text, ext_x, ext_y + y_shift)
        if self.lead_data['FirmName'].values[0] == 'ОБЩИНСКО ПРЕДПРИЯТИЕ СОЦИАЛНО ОБЩЕСТВЕНА ТРАПЕЗАРИ':
            y_shift += 5
            p_text = 'Разходът е направен във връзка с Договор ФС№01-0472/24.04.2020 между АСП и Община Русе, по Целева програма <br/>\"Топъл Обяд у дома в условията на извънредна ситуация 2020\".'
            self.__paragraph_writer__(p_text, ext_x, ext_y + y_shift)
        

    def create_footer(self, ext_x = 10 , ext_y = 280):

        ext_y -=  self.y_offset
        logo = 'app/static/img/backgroung_invoice_GED.png'
        self.canvas.drawInlineImage(logo,*self.coord(ext_x, ext_y, mm),540, 600) 
               
    

    def save(self):    
        self.canvas.save()

    def upload_to_db(self):
        contractor_id= Contractor.query.filter(Contractor.acc_411 == self.lead_data['fullcode'].values[0]).first().id
        if contractor_id is None:
            print(f'Error from invoce to db - missing contractor with acc 411 :{self.lead_data.fullcode.values[0]}')

        invoice_group_id = InvoiceGroup.query.filter(InvoiceGroup.name == self.lead_data['RepFullCode'].values[0]).first().id

        if invoice_group_id is None:
            print(f'Error from invoce to db - missing invouce group with name :{self.lead_data.RepFullCode.values[0]}')

        old_invoice = Invoice.query.filter(Invoice.id == int(self.lead_data['DocNumber'].values[0])).all()

        creation_date = self.lead_data['StockDate'].values[0].replace('.','/')
        creation_date = dt.datetime.strptime(creation_date, '%d/%m/%Y')

        maturity_date = self.lead_data['PayDate'].values[0].replace('.','/')
        maturity_date = dt.datetime.strptime(maturity_date, '%d/%m/%Y')

        ref_file_name = self.lead_data['RepFileName'].values[0].replace('"',' ')
        if len(old_invoice) == 0:

            invoice = Invoice(id = int(self.lead_data['DocNumber'].values[0]),
                                contractor_id = contractor_id,
                                total_qty = Decimal(self.power.Quantity.values[0]),
                                total_sum = Decimal(self.total_sum),
                                grid_sum = Decimal(self.grid.ItemSuma.values[0]),
                                zko_sum = Decimal(self.zko.ItemSuma.values[0]),
                                akciz_sum = Decimal(self.akciz.ItemSuma.values[0]),
                                additional_tax_sum = 0,
                                ref_file_name = ref_file_name,
                                easypay_num = int(self.lead_data['easy_pay_num'].values[0]),
                                easypay_name = self.lead_data['easy_pay_name'].values[0],
                                is_easypay = self.lead_data['EasyPay'].values[0],
                                creation_date = creation_date,
                                maturity_date = maturity_date,
                                price = Decimal(self.power.PriceLev.values[0]),
                                invoice_group_id = invoice_group_id
                            )
            invoice.save()
        else:
            invoice_dict = {'id': int(self.lead_data['DocNumber'].values[0]),
                                'contractor_id' : contractor_id,
                                'total_qty' : Decimal(self.power.Quantity.values[0]),
                                'total_sum' : Decimal(self.total_sum),
                                'grid_sum' : Decimal(self.grid.ItemSuma.values[0]),
                                'zko_sum' : Decimal(self.zko.ItemSuma.values[0]),
                                'akciz_sum' : Decimal(self.akciz.ItemSuma.values[0]),
                                'additional_tax_sum' : 0,
                                'ref_file_name' : ref_file_name,
                                'easypay_num' : int(self.lead_data['easy_pay_num'].values[0]),
                                'easypay_name' : self.lead_data['easy_pay_name'].values[0],
                                'is_easypay' : self.lead_data['EasyPay'].values[0],
                                'creation_date' : creation_date,
                                'maturity_date' : maturity_date,
                                'price' : Decimal(self.power.PriceLev.values[0]),
                                'invoice_group_id' : invoice_group_id
                        }
            old_invoice[0].update(invoice_dict)
            

        
    
def fonts_init():

    pdfmetrics.registerFont(TTFont('DejaVuSerif', 'app/static/fonts/DejaVuSerif.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVuSerifBold', 'app/static/fonts/DejaVuSerifBold.ttf'))
    pdfmetrics.registerFont(TTFont('Art', 'app/static/fonts/Lemon-Tuesday.ttf'))

    addMapping('DejaVuSerif', 0, 0, 'DejaVuSerif') #normal
    # addMapping('Vera', 0, 1, 'Vera-Italic') #italic
    addMapping('DejaVuSerif', 1, 0, 'DejaVuSerifBold') #bold

def ref_num_injector(raw_df):

    filename = os.path.join(os.path.join(app.root_path, app.config['INV_REFS_PATH']) , raw_df.iloc[0].RepFileName)
    try:
        wb = load_workbook(filename = filename)
    except Exception as e:         
        print(f'{e}  \n Exception at row --->{print(sys.exc_info()[2].tb_lineno)}')
    else:
        ws = wb.active
        text = 'СПРАВКА КЪМ ФАКТУРА №'    
        inv_num = raw_df.iloc[0].DocNumber
        final_text = f'{text} {inv_num}'
        ws['A4'].value = final_text
        ws.merge_cells('A4:J4')        
        ws['A4'].alignment = Alignment(wrap_text=True,horizontal='center')
        wb.save(filename)
        
def create_invoices(raw_df):


    fonts_init()
    par_styler = ParagraphStyler()    
    sum_writer = SumTextWriter()
    
    grouped_dict = dict(tuple(raw_df.groupby('DocNumber')))    
    for key in grouped_dict.keys():        
        print(f'{key}')        
        filename = f'{key}.pdf'
        df = grouped_dict[key]
        par_styler.reset()
        invoice = InvoiceCreator(filename, df, par_styler, sum_writer)  
        invoice.create_footer()      
        invoice.create_barcode()
        invoice.create_logo()
        invoice.create_invoice_number()
        invoice.create_bank_data()
        invoice.create_seller_data()
        invoice.create_buyer_data()
        invoice.create_table()
        invoice.create_payment_summary()
        invoice.create_signature_fields()
        invoice.create_notes()
        invoice.save()
        # invoice.upload_to_db()
        ref_num_injector(df)
        invoice.upload_to_db()


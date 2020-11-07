import sys, pytz, datetime as dt
import pandas as pd
import os
import xlrd
import time,re
from decimal import Decimal,ROUND_HALF_UP

from flask import flash
from app.models import *

import xlsxwriter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import openpyxl
import calendar

MONEY_FINAL_ROUND = 2
MONEY_ROUND = 6
ENERGY_ROUND = 3
ENERGY_ROUND_MW = 6
INV_REFS_PATH = 'app/static/inv_ref_files'
INTEGRA_INDIVIDUAL_PATH = 'app/static/integra_individual_files' 
INTEGRA_FOR_UPLOAD_PATH = 'app/static/integra_for_upload' 
PDF_INVOICES_PATH = 'app/static/created_pdf_invoices'

GOODES_CODE = {'Сума за енергия':'304-1', 'Мрежови услуги (лв.)':'498-56','Задължение към обществото':'459-2','Акциз':'456-1'}
PRICES = {'304-1':'price', '498-56':'Мрежови услуги (лв.)','459-2':'zko','456-1':'akciz'}


COL_NAMES = ['№', 'Обект (ИТН №)',  'Адрес', 'Потребление (kWh)',
            'Сума за енергия', 'Задължение към обществото', 'Акциз',
            'Мрежови услуги (лв.)', 'Обща сума (без ДДС)']

def coord_from_string(str_addr):
    xy = coordinate_from_string(str_addr)
    col = column_index_from_string(xy[0]) 
    row = xy[1]
    return (col, row)

def make_header(ws, data, row_num):
    
    fat_cell_border = Border(left=Side(border_style='thick', color='FF000000'),
                     right=Side(border_style='thick', color='FF000000'),
                     top=Side(border_style='thick', color='FF000000'),
                     bottom=Side(border_style='thick', color='FF000000'))
    
    curr_col_idx = coord_from_string(f"""A{row_num}""")[0]
    curr_row_idx = row_num
    for curr_data in data:
        ws.cell(row=curr_row_idx, column=curr_col_idx, value=curr_data)
        ws.cell(row=curr_row_idx, column=curr_col_idx).alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        ws.cell(row=curr_row_idx, column=curr_col_idx).border = fat_cell_border
        ws.cell(row=curr_row_idx, column=curr_col_idx).font =  Font(size=12, color='000000', bold=True, italic=False)
        ws.merge_cells(start_row = curr_row_idx, start_column = curr_col_idx, end_row = curr_row_idx + 1, end_column = curr_col_idx)
        curr_col_idx += 1

def generate_ref_excel(df, df_grid, invoice_start_date, invoice_end_date, period_start_date, period_end_date, is_second = False):        
   
    # contractor = df['contractor_name'].iloc[0]
    contractor = df['invoice_group_description'].iloc[0]    
    # period = f'{calendar.month_name[period_end_date.month]}/{period_end_date.year}'
    period = f'{period_end_date.month}/{period_end_date.year}'
    
    file_name =f'{period_end_date.year}-{period_end_date.month}_{df.iloc[0].invoice_group_description}_{df.iloc[0].invoice_group_name}_invoice_reference.xlsx' 
    
    writer = pd.ExcelWriter(f'{INV_REFS_PATH}/{file_name}', engine='xlsxwriter')
    src_df = pd.read_excel('app/static/uploads/src_dete.xlsx', header=None) if is_second else pd.read_excel('app/static/uploads/src.xlsx', header=None)
    src_df.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
    df_grid.to_excel(writer, sheet_name = 'мрежови услуги')
    writer.close()

    wb = load_workbook(filename = f'{INV_REFS_PATH}/{file_name}')

    ws = wb.active
    
    thin_cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))                    

    navy_blue_fill = PatternFill("solid", fgColor="103ca2")
    ligh_blue_fill = PatternFill("solid", fgColor="9bc2e6")
    pink_fill = PatternFill("solid", fgColor="FCE4D6")
    d_gray_fill = PatternFill("solid", fgColor="595959")
    l_gray_fill = PatternFill("solid", fgColor="D9D9D9")    

    img = openpyxl.drawing.image.Image('app/static/uploads/dete2.png') if is_second else openpyxl.drawing.image.Image('app/static/uploads/Grand_Energy_.png')
    img.anchor = 'I1'
    ws.add_image(img)


    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 38
    # ws.column_dimensions['C'].width = 20
    # ws.column_dimensions['D'].width = 38
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 20  
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20
    ws.merge_cells('A11:D11')

    if is_second:
        for c in ws.iter_cols(1, 7, 11 , 11):
            c[0].fill = navy_blue_fill 

        for c in ws.iter_cols(1, 7, 16 , 17):
            c[0].fill = ligh_blue_fill

        for c in ws.iter_cols(1, 7, 18 , 19):
            c[0].fill = ligh_blue_fill

        for c in ws.iter_cols(1, 7, 20 , 21):
            c[0].fill = navy_blue_fill

        ws['A3'].font =  Font(size=12, color='000000', bold=True, italic=False) 
        # ws['A4'].alignment = Alignment(wrap_text=True,horizontal='center')
        ws['A7'].value = f"""КЛИЕНТ: {contractor}"""
        ws['A7'].font =  Font(size=12, color='000000', bold=True, italic=False)

        ws['A27'].font =  Font(size=12, color='000000', bold=True, italic=False) 
        ws['A9'].value = f"""ПЕРИОД: {period}"""
        ws['A9'].font =  Font(size=12, color='000000', bold=True, italic=False) 
        ws['A11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['A11'].alignment = Alignment(wrap_text=True,horizontal='center')
        ws['E11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['E11'].alignment = Alignment(horizontal='center')
        ws['F11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['F11'].alignment = Alignment(horizontal='center')
        ws['G11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['G11'].alignment = Alignment(horizontal='center')
        ws['A20'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 

        total_consumption = round(df['Потребление (kWh)'].sum() /1000 ,ENERGY_ROUND_MW)
        # print(f'{total_consumption}')
        ws['E12'].value = total_consumption 
        ws['E12'].number_format = '### ### ###.00000' if total_consumption != 0 else '0'

        total_value = df['Сума за енергия'].sum()
        ws['F12'].value = round(((total_value/total_consumption)), MONEY_ROUND) if(total_consumption != 0) else 0
        ws['F12'].number_format = '### ### ##0.00 лв.'

        ws['G12'].value = ws['F12'].value * ws['E12'].value
        ws['G12'].number_format = '### ### ##0.00 лв.'

        grid_services = df['Мрежови услуги (лв.)'].sum()
        ws['E13'].value = total_consumption if grid_services > 0 else ''
        ws['E13'].number_format = '# ##.00000' if total_consumption != 0 else '0'
        ws['E14'].value = total_consumption 
        ws['E14'].number_format = '# ###.00000' if total_consumption != 0 else '0'
        ws['E15'].value = total_consumption 
        ws['E15'].number_format = '# ###.00000' if total_consumption != 0 else '0'
        ws['E16'].value = total_consumption 
        ws['E16'].number_format = '# ###.00000' if total_consumption != 0 else '0'

        
        ws['G13'].value = round(grid_services, MONEY_ROUND)
        ws['G13'].number_format = '### ### ##0.00 лв.'

        zko = df['Задължение към обществото'].sum()
        ws['G14'].value = round(zko, MONEY_ROUND)
        ws['G14'].number_format = '### ### ##0.00 лв.'

        akciz = df['Акциз'].sum()
        ws['G15'].value = round(akciz, MONEY_ROUND)
        ws['G15'].number_format = '### ### ##0.00 лв.'

        ws['G16'].value = round((ws['G12'].value + ws['G13'].value + ws['G14'].value + ws['G15'].value),MONEY_FINAL_ROUND) 
        ws['G16'].number_format = '### ### ##0.00 лв.'

        ws['G18'].value = round((ws['G16'].value * Decimal('0.2')),MONEY_FINAL_ROUND)
        ws['G18'].number_format = '### ### ##0.00 лв.'

        ws['G20'].value = ws['G16'].value + ws['G18'].value
        ws['G20'].number_format = '### ### ##0.00 лв.'
        ws['G20'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 

        # ws['F14'].value = df.iloc[0].zko * 1000 
        # ws['F14'].number_format = '# ##0.00'

        # ws['F15'].value = df.iloc[0].akciz * 1000 
        # ws['F15'].number_format = '# ##0.00'

    else:
        for c in ws.iter_cols(1, 7, 11 , 11):
            c[0].fill = d_gray_fill

        for c in ws.iter_cols(1, 7, 17 , 17):
            c[0].fill = l_gray_fill

        for c in ws.iter_cols(1, 7, 19 , 19):
            c[0].fill = pink_fill

        for c in ws.iter_cols(1, 7, 21 , 21):
            c[0].fill = d_gray_fill

        # ws.merge_cells('A4:J4') 

        ws['A4'].font =  Font(size=12, color='000000', bold=True, italic=False) 
        # ws['A4'].alignment = Alignment(wrap_text=True,horizontal='center')
        ws['A7'].value = f"""КЛИЕНТ: {contractor}"""
        ws['A7'].font =  Font(size=12, color='000000', bold=True, italic=False)

        ws['A27'].font =  Font(size=12, color='000000', bold=True, italic=False) 
        ws['A9'].value = f"""ПЕРИОД: {period}"""
        ws['A9'].font =  Font(size=12, color='000000', bold=True, italic=False) 
        ws['A11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['A11'].alignment = Alignment(wrap_text=True,horizontal='center')
        ws['E11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['E11'].alignment = Alignment(horizontal='center')
        ws['F11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['F11'].alignment = Alignment(horizontal='center')
        ws['G11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
        ws['G11'].alignment = Alignment(horizontal='center')
        ws['A21'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 


        total_consumption = df['Потребление (kWh)'].sum() / 1000
        ws['E12'].value = total_consumption 
        ws['E12'].number_format = '### ### ##0.00000' if total_consumption != 0 else '0'
        # # ws['E12'].value = "{:.2f}".format(total_consumption)

        total_value = df['Сума за енергия'].sum()
        ws['F12'].value = round(((total_value/total_consumption)), MONEY_ROUND) if(total_consumption != 0) else 0
        ws['F12'].number_format = '### ### ##0.00 лв.'
        # ws['F12'].number_format = '### ### ##0.00000 лв.'

        ws['G12'].value = round((ws['F12'].value * ws['E12'].value), MONEY_FINAL_ROUND)
        ws['G12'].number_format = '### ### ##0.00 лв.'

        ws['E13'].value = total_consumption 
        ws['E13'].number_format = '# ##0.00000' if total_consumption != 0 else '0'
        ws['E14'].value = total_consumption 
        ws['E14'].number_format = '# ##0.00000' if total_consumption != 0 else '0'
        ws['E15'].value = total_consumption 
        ws['E15'].number_format = '# ##0.00000' if total_consumption != 0 else '0'
        ws['E16'].value = total_consumption 
        ws['E16'].number_format = '# ##0.00000' if total_consumption != 0 else '0'

        grid_services = df['Мрежови услуги (лв.)'].sum()
        ws['G13'].value = round(grid_services, MONEY_FINAL_ROUND)
        ws['G13'].number_format = '### ### ##0.00 лв.'

        zko = df['Задължение към обществото'].sum()
        ws['G14'].value = round(zko, MONEY_FINAL_ROUND)
        ws['G14'].number_format = '### ### ##0.00 лв.'

        akciz = df['Акциз'].sum()
        ws['G15'].value = round(akciz, MONEY_FINAL_ROUND)
        ws['G15'].number_format = '### ### ##0.00 лв.'

        ws['G16'].value = 0
        ws['G16'].number_format = '### ### ##0.00 лв.'

        ws['G17'].value = round((ws['G12'].value + ws['G13'].value + ws['G14'].value + ws['G15'].value + ws['G16'].value),MONEY_FINAL_ROUND)
        ws['G17'].number_format = '### ### ##0.00 лв.'

        ws['G19'].value = round((ws['G17'].value * Decimal('0.2')),MONEY_FINAL_ROUND)
        ws['G19'].number_format = '### ### ##0.00 лв.'

        ws['G21'].value = ws['G17'].value + ws['G19'].value
        ws['G21'].number_format = '### ### ##0.00 лв.'
        ws['G21'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 

        # ws['F14'].value = ZKO #round((ZKO * Decimal('1000')),MONEY_ROUND)
        # ws['F14'].number_format = '# ##0.00'

        # ws['F15'].value = AKCIZ #round((AKCIZ * Decimal('1000')),MONEY_ROUND)
        # ws['F15'].number_format = '# ##0.00'

    ws['F14'].value = df.iloc[0].zko * 1000 
    ws['F14'].number_format = '# ##0.00'

    ws['F15'].value = df.iloc[0].akciz * 1000 
    ws['F15'].number_format = '# ##0.00'

    final_df = df[['№', 'Обект (ИТН №)', 'Адрес', 'Потребление (kWh)','Сума за енергия','Задължение към обществото', 'Мрежови услуги (лв.)','Акциз']].copy()
    
    final_df['Мрежови услуги (лв.)'] = final_df['Мрежови услуги (лв.)'].apply(lambda x: 0 if x is None else x)
    # final_df.drop(columns = ['Мрежови услуги (лв.)'], inplace = True)
    # final_df.rename(columns = {'Мрежови услуги (лв.)_':'Мрежови услуги (лв.)'}, inplace = True)
    final_df['Обща сума (без ДДС)'] = final_df['Сума за енергия'] + final_df['Акциз'] + final_df['Задължение към обществото'] + final_df['Мрежови услуги (лв.)']

    rows = dataframe_to_rows(final_df,index=False)
    itn_count = final_df.shape[0]
    ws['A10'].value = f"""БРОЙ ОБЕКТИ: {itn_count}"""
    ws['A10'].font =  Font(size=12, color='000000', bold=True, italic=False) 
    
    for r_idx, row in enumerate(rows, 29):

        for c_idx, value in enumerate(row, 1):

            val = value
            if(c_idx == 4):
                ws.cell(row=r_idx, column=c_idx).number_format = '### ### ##0.000'
            elif(c_idx > 4):            
                ws.cell(row=r_idx, column=c_idx).number_format = '### ### ##0.00 лв.'       

            ws.cell(row=r_idx, column=c_idx, value=val).border = thin_cell_border
            ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            ws.cell(row=r_idx, column=c_idx).font =  Font(size=12, color='000000', bold=False, italic=False)

    col_names = list(final_df.columns)
    # col_names.append('Обща сума (без ДДС)')
    make_header(ws, col_names, 28)
    l_row = len(ws['A'])

    ws.print_area = f"""A1:J{l_row}"""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.print_title_rows = '28:29'
    wb.save(f'{INV_REFS_PATH}/{file_name}')
    return file_name
   
    

def generate_num_and_name(first_digit, num_411, inv_group, name):
    
    prefix = num_411.rsplit('-',1)[1]    
    suffix = inv_group.rsplit('_',1)[1]
    prefix_zeroes = (5-len(prefix)) * '0'
    suffix_zeroes = (3-len(suffix)) * '0'
    num_str = f'{first_digit}{prefix_zeroes}{prefix}{suffix_zeroes}{suffix}'
    # name_str =  name + ' ' + str(suffix) if suffix != '0' else name
    name_str =  name + ' ' + str(suffix) if suffix != '1' else name
            
    return (num_str, name_str)    

def generate_integra_file(df, start_date, end_date, ref_file_name):
    
    df = df.fillna(Decimal('0'))
   
    inv_group_name = df.iloc[0]['invoice_group_name'] 
    
    curr_contract = db.session.query(Contract).join(SubContract).join(InvoiceGroup).filter(InvoiceGroup.name == inv_group_name ).filter(SubContract.start_date <= start_date, SubContract.end_date > start_date).first()
  
    try:
        df['price'] = (df['Сума за енергия'].sum()) / (df['Потребление (kWh)'].sum())
    except:
        df['price'] = Decimal('0')


    for_invoice_df = df[['Потребление (kWh)','Сума за енергия','Мрежови услуги (лв.)','Задължение към обществото','Акциз']].sum()
   
    for_invoice_df = for_invoice_df.to_frame().T   
    for_invoice_df['inv_group']=df.iloc[0]['invoice_group_name'] 
    for_invoice_df['Получател']=df.iloc[0]['contractor_name']    
    for_invoice_df['сметка 411']=df['invoice_group_name'].iloc[0].split('_')[0]

    last_month_date = end_date.replace(day = calendar.monthrange(end_date.year, end_date.month)[1])
    for_invoice_df['Дата на издаване'] = last_month_date.strftime('%d/%m/%Y')
    maturity_date = (dt.date.today() + pd.offsets.BDay(curr_contract.maturity_interval)).strftime('%d/%m/%Y') if curr_contract.maturity_interval <= 15 else (dt.date.today() + dt.timedelta(days = curr_contract.maturity_interval)).strftime('%d/%m/%Y')
    
    for_invoice_df['Падеж'] = maturity_date
    reason_date_str = last_month_date.strftime('%m.%Y') 
    for_invoice_df['Основание'] = f' за м.{reason_date_str}г.'
    
    epay_code, epay_name =  generate_num_and_name(1, for_invoice_df.iloc[0]['сметка 411'], for_invoice_df.iloc[0]['inv_group'],for_invoice_df.iloc[0]['Получател'])
    for_invoice_df['easy_pay_num'] = epay_code
    for_invoice_df['easy_pay_name'] = epay_name
    for_invoice_df = pd.melt(for_invoice_df, id_vars=['Потребление (kWh)', 'inv_group','Получател','сметка 411','Дата на издаване','Падеж','Основание','easy_pay_num','easy_pay_name'],var_name = 'Код на стоката',value_name = 'Стойност без ДДС')
    
    for_invoice_df['Стойност без ДДС'] = for_invoice_df['Стойност без ДДС'].apply(lambda x: round(Decimal(x) ,2))

    # print(f'{for_invoice_df}')
    for_invoice_df['Код на стоката'] = for_invoice_df['Код на стоката'].apply(lambda x: GOODES_CODE[x])
    for_invoice_df['Основание'] = for_invoice_df.apply(lambda x: x['Основание'] if x['Код на стоката'] == '304-1' else '', axis = 1)
    for_invoice_df['Количество'] = for_invoice_df.apply(lambda x: Decimal(str(x['Потребление (kWh)'])) / Decimal('1000') if x['Код на стоката'] != '498-56' else 1, axis = 1)
    
    for_invoice_df['Цена без ДДС'] = for_invoice_df.apply(lambda x: df.iloc[0][PRICES[x['Код на стоката']]] * 1000 if x['Код на стоката'] != '498-56' else x['Стойност без ДДС'], axis = 1)
    
    for_invoice_df['номер на фактура'] = ''
    for_invoice_df['Дименсия на количество'] = ''
    for_invoice_df['ЕИК'] = ''
    for_invoice_df['Валутен курс'] = 1
    for_invoice_df['ТИП на сделката по ДДС'] = 256
    for_invoice_df['ДДС %'] = 20
    for_invoice_df['ДДС'] = for_invoice_df['Стойност без ДДС'].apply(lambda x: round(Decimal(x) * Decimal('0.2') ,2))
    for_invoice_df['Крайна сума'] = for_invoice_df['Стойност без ДДС'].apply(lambda x: round(x * Decimal('1.2') ,2))
    for_invoice_df['email'] = curr_contract.contractor.email
    for_invoice_df['Код на валутата'] = 'лв'

    date_str = last_month_date.strftime('%Y-%m')
    file_name =f'{date_str}_{df.iloc[0].invoice_group_description}_{df.iloc[0].invoice_group_name}_integra.xlsx' 
    for_invoice_df['file_name'] = ref_file_name    

    for_invoice_df = for_invoice_df[['Получател','сметка 411','ЕИК','номер на фактура','Дата на издаване','Падеж', 'Основание','Код на стоката', 'Количество', 'Дименсия на количество', 
                                    'Цена без ДДС', 'Код на валутата', 'Валутен курс','Стойност без ДДС','ТИП на сделката по ДДС','ДДС %',
                                    'ДДС','Крайна сума', 'inv_group', 'email', 'file_name','easy_pay_num', 'easy_pay_name']]
    
    for_invoice_df.insert(loc=0, column = '№ по ред', value = 1 )
    # for_invoice_df.to_excel(INTEGRA_INDIVIDUAL_PATH + '/' + file_name,index = False)
    for_invoice_df.to_excel(os.path.join(INTEGRA_INDIVIDUAL_PATH, file_name),index = False)
    
    

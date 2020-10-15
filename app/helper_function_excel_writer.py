import sys, pytz, datetime as dt
import pandas as pd
import os
import xlrd
import time,re
from decimal import Decimal,ROUND_HALF_UP

from flask import flash
from app.models import *

import xlsxwriter
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import openpyxl
import calendar

MONEY_ROUND = 2
ENERGY_ROUND = 3
ENERGY_ROUND_MW = 6

# COL_NAMES = ['№', 'Обект (ИТН №)', 'Град/Село', 'Адрес', 'Потребление (kWh)',
#             'Сума за енергия', 'Задължение към обществото', 'Акциз',
#             'Мрежови услуги (лв.)', 'Обща сума (без ДДС)']
COL_NAMES = ['№', 'Обект (ИТН №)',  'Адрес', 'Потребление (kWh)',
            'Сума за енергия', 'Задължение към обществото', 'Акциз',
            'Мрежови услуги (лв.)', 'Обща сума (без ДДС)']

def coord_from_string(str_addr):
    xy = coordinate_from_string(str_addr)
    col = column_index_from_string(xy[0]) 
    row = xy[1]
    return (col, row)

def make_header(ws, data, row_num):
    
    fat_cell_border = Border(left=Side(border_style='thick', color='FF000000'),
                     right=Side(border_style='thick', color='FF000000'),
                     top=Side(border_style='thick', color='FF000000'),
                     bottom=Side(border_style='thick', color='FF000000'))
    
    curr_col_idx = coord_from_string(f"""A{row_num}""")[0]
    curr_row_idx = row_num
    for curr_data in data:
        ws.cell(row=curr_row_idx, column=curr_col_idx, value=curr_data)
        ws.cell(row=curr_row_idx, column=curr_col_idx).alignment = Alignment(wrap_text=True,horizontal='center',vertical='center')
        ws.cell(row=curr_row_idx, column=curr_col_idx).border = fat_cell_border
        ws.cell(row=curr_row_idx, column=curr_col_idx).font =  Font(size=12, color='000000', bold=True, italic=False)
        ws.merge_cells(start_row = curr_row_idx, start_column = curr_col_idx, end_row = curr_row_idx + 1, end_column = curr_col_idx)
        curr_col_idx += 1


def generate_excel(df, df_grid, invoice_start_date, invoice_end_date, period_start_date, period_end_date, time_zone):

    dest_folder_path = 'app/static/generated_excel_files'

   
    contractor = df['contractor_name'].iloc[0]
    # print(f'contractor --> {contractor}')
    period = f'{calendar.month_name[period_end_date.month]}/{period_end_date.year}'
    # print(f'period --> {calendar.month_name[period_start_date.month]}/{period_start_date.year}')
    file_name =f'{period_end_date.year}-{period_end_date.month}_{df.iloc[0].invoice_group_description}%{df.iloc[0].invoice_group_name}%invoice_reference.xlsx' 
    #  print(f'file_name --> {file_name}')    
    
    writer = pd.ExcelWriter(f'{dest_folder_path}/{file_name}', engine='xlsxwriter')
    src_df = pd.read_excel('app/static/uploads/src_dete.xlsx', header=None)
    src_df.to_excel(writer, sheet_name="Sheet1", index=False, header=False)
    df_grid.to_excel(writer, sheet_name = 'мрежови услуги')
    writer.close()

    wb = load_workbook(filename = f'{dest_folder_path}/{file_name}')

    ws = wb.active
    
    thin_cell_border = Border(left=Side(border_style='thin', color='FF000000'),
                     right=Side(border_style='thin', color='FF000000'),
                     top=Side(border_style='thin', color='FF000000'),
                     bottom=Side(border_style='thin', color='FF000000'))                    

    navy_blue_fill = PatternFill("solid", fgColor="103ca2")
    ligh_blue_fill = PatternFill("solid", fgColor="9bc2e6")    

    img = openpyxl.drawing.image.Image('app/static/uploads/dete2.png')
    img.anchor = 'I1'
    ws.add_image(img)


    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 38
    # ws.column_dimensions['C'].width = 20
    # ws.column_dimensions['D'].width = 38
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 20  
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 20
    ws.column_dimensions['J'].width = 20

    for c in ws.iter_cols(1, 7, 11 , 11):
        c[0].fill = navy_blue_fill

    for c in ws.iter_cols(1, 7, 16 , 17):
        c[0].fill = ligh_blue_fill

    for c in ws.iter_cols(1, 7, 18 , 19):
        c[0].fill = ligh_blue_fill

    for c in ws.iter_cols(1, 7, 20 , 21):
        c[0].fill = navy_blue_fill
   

    ws.merge_cells('A11:D11')
    # ws.merge_cells('A4:J4')

    ws['A3'].font =  Font(size=12, color='000000', bold=True, italic=False) 
    # ws['A4'].alignment = Alignment(wrap_text=True,horizontal='center')
    ws['A7'].value = f"""КЛИЕНТ: {contractor}"""
    ws['A7'].font =  Font(size=12, color='000000', bold=True, italic=False)

    ws['A27'].font =  Font(size=12, color='000000', bold=True, italic=False) 
    ws['A9'].value = f"""ПЕРИОД: {period}"""
    ws['A9'].font =  Font(size=12, color='000000', bold=True, italic=False) 
    ws['A11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
    ws['A11'].alignment = Alignment(wrap_text=True,horizontal='center')
    ws['E11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
    ws['E11'].alignment = Alignment(horizontal='center')
    ws['F11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
    ws['F11'].alignment = Alignment(horizontal='center')
    ws['G11'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 
    ws['G11'].alignment = Alignment(horizontal='center')
    ws['A20'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 

    total_consumption = round(df['Потребление (kWh)'].sum() /1000 ,ENERGY_ROUND_MW)
    print(f'{total_consumption}')
    ws['E12'].value = total_consumption 
    ws['E12'].number_format = '### ### ###.00000' if total_consumption != 0 else '0'

    total_value = df['Сума за енергия'].sum()
    ws['F12'].value = round(((total_value/total_consumption)), MONEY_ROUND) if(total_consumption != 0) else 0
    ws['F12'].number_format = '### ### ##0.00 лв.'

    ws['G12'].value = ws['F12'].value * ws['E12'].value
    ws['G12'].number_format = '### ### ##0.00 лв.'

    grid_services = df['Мрежови услуги (лв.)'].sum()
    ws['E13'].value = total_consumption if grid_services > 0 else ''
    ws['E13'].number_format = '# ###.00000' if total_consumption != 0 else '0'
    ws['E14'].value = total_consumption 
    ws['E14'].number_format = '# ###.00000' if total_consumption != 0 else '0'
    ws['E15'].value = total_consumption 
    ws['E15'].number_format = '# ###.00000' if total_consumption != 0 else '0'
    ws['E16'].value = total_consumption 
    ws['E16'].number_format = '# ###.00000' if total_consumption != 0 else '0'

    
    ws['G13'].value = round(grid_services, MONEY_ROUND)
    ws['G13'].number_format = '### ### ##0.00 лв.'

    zko = df['Задължение към обществото'].sum()
    ws['G14'].value = round(zko, MONEY_ROUND)
    ws['G14'].number_format = '### ### ##0.00 лв.'

    akciz = df['Акциз'].sum()
    ws['G15'].value = round(akciz, MONEY_ROUND)
    ws['G15'].number_format = '### ### ##0.00 лв.'

    ws['G16'].value = ws['G12'].value + ws['G13'].value + ws['G14'].value + ws['G15'].value 
    ws['G16'].number_format = '### ### ##0.00 лв.'

    ws['G18'].value = round((ws['G16'].value * Decimal('0.2')),MONEY_ROUND)
    ws['G18'].number_format = '### ### ##0.00 лв.'

    ws['G20'].value = round((ws['G16'].value + ws['G18'].value),MONEY_ROUND)
    ws['G20'].number_format = '### ### ##0.00 лв.'
    ws['G20'].font =  Font(size=12, color='FFFFFF', bold=True, italic=False) 

    ws['F14'].value = df.iloc[0].zko * 1000 
    ws['F14'].number_format = '# ##0.00'

    ws['F15'].value = df.iloc[0].akciz * 1000 
    ws['F15'].number_format = '# ##0.00'

    df = df[['№', 'Обект (ИТН №)', 'Адрес', 'Потребление (kWh)','Сума за енергия','Акциз', 'Задължение към обществото','Мрежови услуги (лв.)']]
    df['Мрежови услуги (лв.)'] = df['Мрежови услуги (лв.)'].apply(lambda x: 0 if x is None else x)
    df['Обща сума (без ДДС)'] = df['Сума за енергия'] + df['Акциз'] + df['Задължение към обществото'] + df['Мрежови услуги (лв.)']

    rows = dataframe_to_rows(df,index=False)
    itn_count = df.shape[0]
    ws['A10'].value = f"""БРОЙ ОБЕКТИ: {itn_count}"""
    ws['A10'].font =  Font(size=12, color='000000', bold=True, italic=False) 


    for r_idx, row in enumerate(rows, 29):

        for c_idx, value in enumerate(row, 1):

            val = value
            if(c_idx == 4):
                ws.cell(row=r_idx, column=c_idx).number_format = '### ### ##0.000'
            elif(c_idx > 4):            
                ws.cell(row=r_idx, column=c_idx).number_format = '### ### ##0.00 лв.'       

            ws.cell(row=r_idx, column=c_idx, value=val).border = thin_cell_border
            ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            ws.cell(row=r_idx, column=c_idx).font =  Font(size=12, color='000000', bold=False, italic=False)

    make_header(ws, COL_NAMES, 28)
    l_row = len(ws['A'])

    ws.print_area = f"""A1:J{l_row}"""
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False
    ws.print_title_rows = '28:29'
    wb.save(f'{dest_folder_path}/{file_name}')
